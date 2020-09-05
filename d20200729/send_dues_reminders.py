# Sends emails based on payment status in spreadsheet.
import smtplib
import sys

import openpyxl

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook("duesRecords.xlsx")
sheet = wb["Sheet1"]

last_col = sheet.max_column
latest_month = sheet.cell(row=1, column=last_col).value

# Check each member's payment status.
unpaid_members = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=last_col).value
    if payment != "paid":
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaid_members[name] = email

# Log in to email account.
smtpObj = smtplib.SMTP("smtp.gmail.com", 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login("bokyeong3659@gmail.com", sys.argv[1])

# Send out reminder emails.
for name, email in unpaid_members.items():
    body = f"""Subject: {latest_month} dues unpaid.\nDear {name},
    \nRecodes show that you have not paid dues for {latest_month}.
    Please make this payment as soon as possible. Thank you!"""
    print(f"Sending email to {email}...")
    sendmailStatus = smtpObj.sendmail("bokyeong3659@gmail.com", email, body)

    if sendmailStatus != {}:
        print(f"There was a problom sending email to {email}:{sendmailStatus}")
smtpObj.quit()
