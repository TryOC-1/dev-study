import os

import openpyxl

for excelFile in os.listdir("."):
    if not excelFile.endswith(".xlsx"):
        continue

    wb = openpyxl.load_workbook(excelFile)
    # Skip non-xlsx files, load the workbook object.
    for sheetName in wb.get_sheet_names():
        # Loop through every sheet in the workbook.
        sheet = wb.get_sheet_by_name(sheetName)

        # Create the CSV filename from the Excel filename and sheet title.
        title = excelFile.split(".")[0]
        sheets = sheet.title

        csvFile_name = f"{title}_{sheets}.csv"
        # Create the csv.writer object for this CSV file.
        with open(csvFile_name, "w") as f:
            # Loop through every row in the sheet.
            for rowNum in range(1, sheet.max_row + 1):
                # Loop through each cell in the row.
                col_list = []
                for colNum in range(1, sheet.max_column + 1):
                    # Append each cell's data to rowData.
                    col_list.append(sheet.cell(row=rowNum, column=colNum).value)
                f.write(",".join(col_list) + "\n")
