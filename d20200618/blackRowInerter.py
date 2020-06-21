import sys

import openpyxl

start_row = int(sys.argv[1])
jump_size = int(sys.argv[2])
old_file = sys.argv[3]

old_wb = openpyxl.load_workbook(old_file)
old_sheet = old_wb.active

wb = openpyxl.Workbook()
wb.sheetnames
sheet = wb.active

for row_obj in old_sheet.rows:
    for cell in row_obj:
        row_index = cell.row
        col_index = cell.column

        if row_index >= start_row:
            row_index += jump_size

        sheet.cell(row=row_index, column=col_index).value = cell.value

wb.save("myProduce.xlsx")
