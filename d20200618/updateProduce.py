# 루프로 모든 행을 차례대로 거친다.
# 행이 마늘, 샐러리 또는 레몬이라면 가격을 고친다.
# 1. 스프레드시트 파일을 연다.
# 2. 각 행에 대해 열 A의 값이 Celery(샐러리), Garlic(마늘), Lemon(레몬)인지 검사한다.
# 3. 위 경우에 해당된다면 열 B의 가격을 고친다.
# 4. 스프레드시트를 새로운 파일에 저장한다.(기존 스프레드시트를 잃지 않기위해)

import openpyxl

wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb.get_sheet_by_name("Sheet")

# The produce types and their updated prices
PRICE_UPDATES = {"Gerilc": 3.07, "Celery": 1.19, "Lemon": 1.27}

# Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save("updateProduceSales.xlsx")
