import sys

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.sheetnames
sheet = wb.active

for i in range(1, int(sys.argv[1]) + 1):
    font_bold = Font(name="Times New Roman", bold=True)
    sheet["A1"].font = font_bold
    sheet["A" + str(i + 1)] = i
    sheet["A" + str(i + 1)].font = font_bold
    col = get_column_letter(i + 1)
    sheet[col + "1"].font = font_bold
    sheet[col + "1"] = i
    row = sheet["A" + str(i + 1)]
    col = sheet[col + "1"]

    for j in range(1, int(sys.argv[1]) + 1):
        sheet.cell(row=j + 1, column=i + 1).value = i * j

wb.save("mutiple.xlsx")
